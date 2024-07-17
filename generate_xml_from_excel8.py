from xml.sax.saxutils import escape
from openpyxl import load_workbook
from unidecode import unidecode

# Función para crear un elemento CDATA
def create_cdata_element(tag, content):
    element = f'<{tag}><![CDATA[{content}]]></{tag}>'
    return element

# Función para mapear la importancia a su correspondiente valor numérico
def map_importance(importance):
    importance = importance.lower()
    if "high" in importance:
        return "3"
    elif "medium" in importance:
        return "2"
    elif "low" in importance:
        return "1"
    else:
        return ""

# Cargar el archivo Excel
wb = load_workbook(filename='Template_Oficial.xlsx', read_only=True)
sheet = wb.active

# Leer nombre del archivo XML del Excel y limpiar/formato el nombre
xml_filename = sheet['I2'].value if sheet['I2'].value else "converted_testcases.xml"
xml_filename = unidecode(xml_filename).replace(' ', '_')  # Eliminar acentos y reemplazar espacios por guiones bajos

# Verificar si el nombre de archivo tiene la extensión .xml
if not xml_filename.endswith('.xml'):
    xml_filename += '.xml'

# Diccionario para almacenar casos de prueba
testcases = {}

# Leer datos del archivo Excel y crear casos de prueba
current_name = None
for row in range(2, sheet.max_row + 1):
    name = sheet[f'A{row}'].value
    importance = sheet[f'B{row}'].value
    summary = sheet[f'C{row}'].value
    preconditions = sheet[f'D{row}'].value
    actions = sheet[f'E{row}'].value
    expected_results = sheet[f'F{row}'].value
    custom_name = sheet[f'G{row}'].value
    custom_value = sheet[f'H{row}'].value
    
    # Validar que el nombre no sea None ni esté vacío
    if name is None or name.strip() == "":
        continue  # Saltar esta fila si el nombre está vacío
    
    # Si hay un nombre nuevo, crear un nuevo testcase
    if name != current_name:
        current_name = name
        testcases[current_name] = {
            "name": current_name,
            "importance": importance if importance else "",
            "summary": summary if summary else "",
            "preconditions": preconditions if preconditions else "",
            "steps": [],
            "custom_fields": []  # Usar una lista para almacenar los custom_fields
        }

    # Agregar actions al testcase actual
    if actions:
        if isinstance(actions, str):  # Verificar si es una cadena
            actions = actions.split('-')  # Convertir a lista si es una cadena separada por <br>
        for action in actions:
            testcases[current_name]["steps"].append({
                "action": action.strip(),
                "expected_result": ""  # Inicialmente vacío, ya que se llenará después
            })
    
    # Agregar expected results al testcase actual
    if expected_results:
        if isinstance(expected_results, str):  # Verificar si es una cadena
            expected_results = expected_results.split('-')  # Convertir a lista si es una cadena separada por <br>
        for idx, result in enumerate(expected_results, start=1):
            if len(testcases[current_name]["steps"]) >= idx:  # Asegurar que haya un step correspondiente
                testcases[current_name]["steps"][idx - 1]["expected_result"] = result.strip()
    
    # Agregar custom fields al testcase actual
    if custom_name and custom_value:
        if isinstance(custom_name, str):  # Verificar si es una cadena
            custom_name = custom_name.split('-')  # Convertir a lista si es una cadena separada por <br>
        if isinstance(custom_value, str):  # Verificar si es una cadena
            custom_value = custom_value.split('-')  # Convertir a lista si es una cadena separada por <br>
        
        # Asegurarse de que tengan la misma longitud o tomar la más corta
        num_custom_fields = min(len(custom_name), len(custom_value))
        for i in range(num_custom_fields):
            testcases[current_name]["custom_fields"].append({
                "name": custom_name[i].strip(),
                "value": custom_value[i].strip()
            })

# Generar XML para todos los casos de prueba recolectados
xml_string = '<?xml version="1.0" encoding="UTF-8"?>\n'
xml_string += '<testcases>\n'
for name, data in testcases.items():
    xml_string += f'    <testcase internalid="" name="{escape(name)}">\n'
    xml_string += f'        <node_order><![CDATA[]]></node_order>\n'
    xml_string += f'        <externalid><![CDATA[]]></externalid>\n'
    xml_string += f'        <version><![CDATA[]]></version>\n'
    xml_string += f'        <summary><![CDATA[{escape(data["summary"])}]]></summary>\n'
    xml_string += f'        <preconditions><![CDATA[{escape(data["preconditions"])}]]></preconditions>\n'
    xml_string += f'        <execution_type><![CDATA[1]]></execution_type>\n'
    xml_string += f'        <importance><![CDATA[{map_importance(data["importance"])}]]></importance>\n'
    
    # Steps
    if data["steps"]:
        xml_string += f'        <steps>\n'
        for index, step in enumerate(data["steps"], start=1):
            xml_string += f'            <step>\n'
            xml_string += f'                <step_number><![CDATA[{index}]]></step_number>\n'
            xml_string += f'                <actions><![CDATA[{escape(step["action"])}]]></actions>\n'
            xml_string += f'                <expectedresults><![CDATA[{escape(step["expected_result"])}]]></expectedresults>\n'
            xml_string += f'                <execution_type><![CDATA[1]]></execution_type>\n'
            xml_string += f'            </step>\n'
        xml_string += f'        </steps>\n'
    
    # Custom Fields
    if data["custom_fields"]:
        xml_string += f'        <custom_fields>\n'
        for cf in data["custom_fields"]:
            xml_string += f'            <custom_field>\n'
            xml_string += f'                <name><![CDATA[{escape(cf["name"])}]]></name>\n'
            xml_string += f'                <value><![CDATA[{escape(cf["value"])}]]></value>\n'
            xml_string += f'            </custom_field>\n'
        xml_string += f'        </custom_fields>\n'
    
    xml_string += f'    </testcase>\n'

xml_string += '</testcases>'

# Guardar el XML en un archivo con la extensión .xml
with open(xml_filename, "w", encoding="utf-8") as xml_file:
    xml_file.write(xml_string)

print(f"Archivo XML generado y guardado como '{xml_filename}'")